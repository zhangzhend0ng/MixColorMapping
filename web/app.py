#!/usr/bin/env python3
"""color-mixer-batch web UI — fast version.

Zero-dependency (stdlib only) local web server. On startup it spawns the
compiled CLI ONCE in --serve mode and keeps it alive, talking to it over
stdin/stdout pipes (one JSON request → one JSON response per line). This
avoids the ~50ms process-spawn cost per query that made the old version slow.

Run:  python3 web/app.py   (or ./web/start.sh)
Then open http://localhost:8008

Requires build/color_match_batch[.exe] to be built first (run build.sh).
"""
import atexit
import io
import json
import os
import socketserver
import subprocess
import sys
import threading
from http.server import BaseHTTPRequestHandler
from urllib.parse import unquote

# openpyxl is the only third-party dependency (for .xlsx read/write). Graceful
# degradation: the CSV/web-edit path works without it; only xlsx upload needs it.
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# When frozen with PyInstaller, resources live under sys._MEIPASS (a temp dir
# PyInstaller extracts at startup). When run from source, they live alongside
# this file. Detect once at import time.
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    HERE = sys._MEIPASS
    ROOT = HERE
    STATIC_DIR = os.path.join(HERE, "static")
else:
    HERE = os.path.dirname(os.path.abspath(__file__))
    ROOT = os.path.dirname(HERE)
    STATIC_DIR = os.path.join(HERE, "static")
DEFAULT_PORT = 8008


def find_cli():
    """Locate the compiled CLI binary.

    Priority:
      1. Next to the frozen exe (dev/rel override) — let users drop a freshly
         built CLI beside the .exe without rebuilding the bundle.
      2. Inside the PyInstaller bundle (_MEIPASS/build/color_match_batch[.exe]).
      3. From source tree (build/color_match_batch[.exe]).
    """
    names = ("color_match_batch.exe", "color_match_batch")
    # 1. beside the frozen exe
    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)
        for name in names:
            p = os.path.join(exe_dir, name)
            if os.path.isfile(p):
                return p
    # 2/3. inside bundle or source tree
    for name in names:
        p = os.path.join(ROOT, "build", name)
        if os.path.isfile(p):
            return p
    return None


# ---------------------------------------------------------------------------
# Persistent CLI worker: one long-lived subprocess, JSON-over-pipes.
# ---------------------------------------------------------------------------

class CliWorker:
    """Manages the --serve subprocess. Thread-safe via a lock per call."""

    def __init__(self, cli_path):
        self.cli_path = cli_path
        self.proc = None
        self._lock = threading.Lock()
        self._start()

    def _start(self):
        self.proc = subprocess.Popen(
            [self.cli_path, "--serve"],
            stdin=subprocess.PIPE, stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            bufsize=1, text=True, encoding="utf-8",
        )
        # Wait for the ready handshake.
        ready = self.proc.stdout.readline()
        if not ready.strip().startswith("{"):
            raise RuntimeError("CLI did not signal ready; got: %r" % ready)

    def query(self, request_dict):
        """Send one request, read one response. Returns parsed JSON dict."""
        line = json.dumps(request_dict, separators=(",", ":")) + "\n"
        with self._lock:
            if self.proc.poll() is not None:
                # process died — restart
                self._start()
            try:
                self.proc.stdin.write(line)
                self.proc.stdin.flush()
                resp = self.proc.stdout.readline()
            except (BrokenPipeError, OSError):
                self._start()
                self.proc.stdin.write(line)
                self.proc.stdin.flush()
                resp = self.proc.stdout.readline()
        if not resp:
            return {"error": "CLI returned empty response (crashed?)"}
        try:
            return json.loads(resp)
        except json.JSONDecodeError as e:
            return {"error": "bad JSON from CLI: %s — %r" % (e, resp[:200])}

    def shutdown(self):
        try:
            if self.proc and self.proc.poll() is None:
                self.proc.stdin.close()
                self.proc.wait(timeout=2)
        except Exception:
            try:
                self.proc.kill()
            except Exception:
                pass


# Global worker — created lazily on first use so the server starts fast even
# if the CLI isn't ready yet (e.g. still building).
_WORKER = None
_WORKER_LOCK = threading.Lock()


def get_worker():
    global _WORKER
    if _WORKER is not None and _WORKER.proc and _WORKER.proc.poll() is None:
        return _WORKER
    with _WORKER_LOCK:
        if _WORKER is not None and _WORKER.proc and _WORKER.proc.poll() is None:
            return _WORKER
        cli = find_cli()
        if not cli:
            return None
        _WORKER = CliWorker(cli)
        return _WORKER


# ---------------------------------------------------------------------------
# Default sample data
# ---------------------------------------------------------------------------

DEFAULT_PALETTE = [
    ["#00FFFF", "Cyan"],    # C
    ["#FF00FF", "Magenta"], # M
    ["#FFFF00", "Yellow"],  # Y
    ["#FFFFFF", "White"],   # W
]
DEFAULT_TARGETS_HEX = [
    ["orange", "#FF8800"], ["teal", "#008888"], ["purple", "#8800FF"],
    ["olive", "#666600"], ["gray", "#808080"], ["pink", "#FFAAAA"],
]


# ---------------------------------------------------------------------------
# XLSX processing: read Lab targets from a sheet, batch-match, write results
# back into new sheets on the same workbook.
# ---------------------------------------------------------------------------

def _hex_fill(hex_str):
    """openpyxl PatternFill from a #rrggbb string (for colored result cells)."""
    h = hex_str.lstrip("#")
    if len(h) >= 6:
        return PatternFill(start_color="FF" + h[:6], end_color="FF" + h[:6], fill_type="solid")
    return None


def process_xlsx(xlsx_bytes, palette_rows, min_percent, exclude_str):
    """Read Lab targets from a sheet, batch-match with the CLI worker, write a
    single combined result sheet '配色结果' with CMYW percentage columns for
    both FS and Prusa recipes side by side, plus a green highlight on the
    better algorithm's cells per row (lower ΔE wins).

    Returns the modified workbook as bytes (xlsx). Raises on error.
    Palette format: [["#hex","label"], ...] — same as the web UI uses.
    Expects a CMYW palette (Cyan/Magenta/Yellow/White) for column mapping;
    palette ids whose hex doesn't match a standard CMYW value are dropped.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
    if not palette_rows or len(palette_rows) < 2:
        raise RuntimeError("palette needs >= 2 colors")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    # ---- find the targets sheet: prefer "Sheet1", else first non-empty sheet ----
    targets_ws = None
    for name in ["Sheet1", "sheet1", "Targets", "目标色"]:
        if name in wb.sheetnames:
            targets_ws = wb[name]
            break
    if targets_ws is None:
        # pick the sheet with the most rows
        targets_ws = max(wb.worksheets, key=lambda ws: ws.max_row)
    if targets_ws.max_row < 2:
        raise RuntimeError("targets sheet '%s' appears empty" % targets_ws.title)

    # ---- parse palette rows → [hex or label, hex] ----
    palette = []
    for row in palette_rows:
        hex_val = None
        if row and row[0] and str(row[0]).startswith("#"):
            hex_val = row[0]
        elif len(row) > 1 and row[1] and str(row[1]).startswith("#"):
            hex_val = row[1]
        if hex_val:
            palette.append(hex_val)

    # ---- read targets (Lab format: 序号, L*, a*, b*) ----
    # Detect columns: find header row, then read L/a/b by header name.
    # Header matching is whitespace/case-insensitive and ignores trailing
    # stars, so "L*", "l*", "L", "Lab_L" all map to the L channel.
    headers = {}
    for col in range(1, targets_ws.max_column + 1):
        val = targets_ws.cell(row=1, column=col).value
        if val is not None:
            # normalized key: lowercase, no spaces, no stars
            key = str(val).strip().lower().replace(" ", "").replace("*", "")
            headers[key] = col

    def find_col(*names):
        for n in names:
            k = n.lower().replace(" ", "").replace("*", "")
            if k in headers:
                return headers[k]
        return None

    col_id = find_col("序号", "id", "no", "index", "#", "编号")
    col_L = find_col("l", "l*", "l*a*b*", "lightness", "lab_l", "明度", "亮度")
    col_a = find_col("a", "a*", "lab_a", "红绿")
    col_b = find_col("b", "b*", "lab_b", "黄蓝")

    if col_L is None or col_a is None or col_b is None:
        # fallback: assume columns B/C/D = L/a/b (A = id)
        col_id, col_L, col_a, col_b = 1, 2, 3, 4

    # collect target rows: {label, L, a, b}
    targets = []
    for r in range(2, targets_ws.max_row + 1):
        L = targets_ws.cell(row=r, column=col_L).value
        a = targets_ws.cell(row=r, column=col_a).value
        b = targets_ws.cell(row=r, column=col_b).value
        label = targets_ws.cell(row=r, column=col_id).value if col_id else r - 1
        if L is None or a is None or b is None:
            continue
        try:
            targets.append({"label": str(label), "L": float(L), "a": float(a), "b": float(b)})
        except (ValueError, TypeError):
            continue

    if not targets:
        raise RuntimeError("no valid Lab target rows found in '%s'" % targets_ws.title)

    # ---- run batch match via the persistent CLI worker ----
    worker = get_worker()
    if worker is None:
        raise RuntimeError("CLI worker not available")

    # CLI expects targets as [["label", L, a, b], ...] in "lab" format
    target_rows = [[t["label"], str(t["L"]), str(t["a"]), str(t["b"])] for t in targets]
    request = {
        "palette": palette_rows,
        "targets": target_rows,
        "format": "lab",
        "rgb_scale": "auto",
        "min_percent": min_percent if min_percent is not None else 0,
        "exclude": exclude_str or "",
    }
    result = worker.query(request)
    if "error" in result and "rows" not in result:
        raise RuntimeError("CLI error: %s" % result["error"])

    rows = result.get("rows", [])
    if not rows:
        raise RuntimeError(
            "CLI returned no result rows (parsed %d target rows from sheet '%s'). "
            "Check that L/a/b columns contain numbers."
            % (len(targets), targets_ws.title)
        )

    # ---- build palette id → CMYW slot mapping ----
    # The CLI returns recipes as palette-index ids ("1/3/4") + weights
    # ("30/50/20"), aligned positionally (ids[i] ↔ weights[i], 1-based).
    # We map each palette id to a fixed C/M/Y/W column by matching its hex
    # against the standard subtractive primaries.
    SLOT_HEX = {
        "C": "00ffff",  # Cyan
        "M": "ff00ff",  # Magenta
        "Y": "ffff00",  # Yellow
        "W": "ffffff",  # White
    }
    id_to_slot = {}
    for pid, hex_ in enumerate(palette, 1):
        norm = hex_.strip().lstrip("#").lower()
        for slot, sh in SLOT_HEX.items():
            if norm == sh:
                id_to_slot[pid] = slot
                break

    def expand_recipe(ids_str, weights_str):
        """Expand "1/3/4" + "30/50/20" → {"C":0,"M":30,"Y":50,"W":20}.
        Unmapped ids (non-standard palette colors) are silently dropped."""
        cm = {"C": 0, "M": 0, "Y": 0, "W": 0}
        if not ids_str or not weights_str:
            return cm
        ids = str(ids_str).split("/")
        wts = str(weights_str).split("/")
        for pid_s, wt_s in zip(ids, wts):
            try:
                pid = int(pid_s)
                wt = int(wt_s)
            except ValueError:
                continue
            slot = id_to_slot.get(pid)
            if slot:
                cm[slot] += wt
        return cm

    # ---- write single combined result sheet ----
    SHEET = "配色结果"
    # remove old result sheets (new format, plus legacy FS_结果/Prusa_结果
    # from earlier runs so re-uploading doesn't leave stale sheets behind).
    for stale in (SHEET, "FS_结果", "Prusa_结果"):
        if stale in wb.sheetnames:
            del wb[stale]
    ws = wb.create_sheet(SHEET)

    # Columns: target info | FS (CMYW% + 预测色 + 预测Lab + ΔE)
    #                       | Prusa (same layout)
    # Lab 压缩成一个单元格 "L/a/b"。不输出 CMYK——那是从 RGB 算出来的派生值，
    # 对实际混色没有参考意义，反而容易误导。
    COLUMNS = [
        "序号", "目标L", "目标a", "目标b", "目标色",
        # FS
        "FS_C%", "FS_M%", "FS_Y%", "FS_W%", "FS预测色",
        "FS预测Lab", "FS_ΔE",
        # Prusa
        "Prusa_C%", "Prusa_M%", "Prusa_Y%", "Prusa_W%", "Prusa预测色",
        "Prusa预测Lab", "Prusa_ΔE",
    ]
    # ΔE column indices (1-based) — for the "best algorithm" highlight.
    FS_DE_COL = 12
    PRUSA_DE_COL = 19

    def fmt_lab(block):
        """预测色 Lab → "L/a/b" 字符串（保留2位小数）。"""
        try:
            return "{}/{}/{}".format(
                round(block.get("L", 0), 2),
                round(block.get("a", 0), 2),
                round(block.get("b_lab", 0), 2),
            )
        except (TypeError, ValueError):
            return ""

    header_font = Font(bold=True)
    for c, title in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=c, value=title).font = header_font

    best_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE",
                            fill_type="solid")  # light green

    # target Lab lookup by label (rows preserve order, but match defensively)
    tgt_lab = {str(t["label"]): (t["L"], t["a"], t["b"]) for t in targets}

    for r_idx, row in enumerate(rows, 2):
        fs = row.get("fs", {})
        prusa = row.get("prusa", {})
        target = row.get("target", {})
        label = row.get("label", "")
        lab = tgt_lab.get(str(label), ("", "", ""))

        # --- target columns (1-5) ---
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=2, value=lab[0])
        ws.cell(row=r_idx, column=3, value=lab[1])
        ws.cell(row=r_idx, column=4, value=lab[2])
        target_hex = target.get("hex", "")
        c5 = ws.cell(row=r_idx, column=5, value=target_hex)
        if target_hex:
            c5.fill = _hex_fill(target_hex)

        # --- FS columns (6-12): CMYW% | 预测色 | 预测Lab | ΔE ---
        fs_cm = expand_recipe(fs.get("ids"), fs.get("weights"))
        ws.cell(row=r_idx, column=6, value=fs_cm["C"])
        ws.cell(row=r_idx, column=7, value=fs_cm["M"])
        ws.cell(row=r_idx, column=8, value=fs_cm["Y"])
        ws.cell(row=r_idx, column=9, value=fs_cm["W"])
        fs_hex = fs.get("hex", "")
        c10 = ws.cell(row=r_idx, column=10, value=fs_hex)
        if fs_hex:
            c10.fill = _hex_fill(fs_hex)
        ws.cell(row=r_idx, column=11, value=fmt_lab(fs))
        ws.cell(row=r_idx, column=12, value=fs.get("delta_e"))

        # --- Prusa columns (13-19): same layout ---
        pr_cm = expand_recipe(prusa.get("ids"), prusa.get("weights"))
        ws.cell(row=r_idx, column=13, value=pr_cm["C"])
        ws.cell(row=r_idx, column=14, value=pr_cm["M"])
        ws.cell(row=r_idx, column=15, value=pr_cm["Y"])
        ws.cell(row=r_idx, column=16, value=pr_cm["W"])
        pr_hex = prusa.get("hex", "")
        c17 = ws.cell(row=r_idx, column=17, value=pr_hex)
        if pr_hex:
            c17.fill = _hex_fill(pr_hex)
        ws.cell(row=r_idx, column=18, value=fmt_lab(prusa))
        ws.cell(row=r_idx, column=19, value=prusa.get("delta_e"))

        # --- highlight only the better algorithm's ΔE cell (lower wins) ---
        fs_de = fs.get("delta_e")
        pr_de = prusa.get("delta_e")
        if isinstance(fs_de, (int, float)) and isinstance(pr_de, (int, float)):
            if fs_de < pr_de:
                ws.cell(row=r_idx, column=FS_DE_COL).fill = best_fill
            elif pr_de < fs_de:
                ws.cell(row=r_idx, column=PRUSA_DE_COL).fill = best_fill
        # (equal or non-numeric → no highlight)

    # auto column width (rough)
    for c in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 10

    # Open the result sheet directly (not the unchanged targets sheet).
    wb.active = wb[SHEET]

    # ---- serialize workbook to bytes ----
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), len(rows)


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass

    def _send(self, code, body=b"", content_type="text/plain", extra_headers=None):
        if isinstance(body, str):
            body = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        if extra_headers:
            for k, v in extra_headers.items():
                self.send_header(k, v)
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, obj, code=200):
        self._send(code, json.dumps(obj), "application/json")

    def do_GET(self):
        path = unquote(self.path.split("?", 1)[0])
        if path in ("/", "/index.html"):
            self._serve_file(os.path.join(STATIC_DIR, "index.html"), "text/html; charset=utf-8")
        elif path == "/api/defaults":
            self._send_json({
                "palette": DEFAULT_PALETTE,
                "targets": DEFAULT_TARGETS_HEX,
                "format": "hex",
                "min_percent": 0,
                "exclude": "",
                "rgb_scale": "auto",
                "cli_found": find_cli() is not None,
            })
        else:
            rel = path.lstrip("/")
            candidate = os.path.join(STATIC_DIR, rel)
            if os.path.isfile(candidate):
                ctype = {"css": "text/css", "js": "application/javascript"}.get(
                    rel.rsplit(".", 1)[-1], "application/octet-stream")
                self._serve_file(candidate, ctype)
            else:
                self._send(404, "not found")

    def _serve_file(self, path, ctype):
        try:
            with open(path, "rb") as f:
                self._send(200, f.read(), ctype)
        except OSError:
            self._send(404, "not found")

    def do_POST(self):
        path = unquote(self.path)
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b""
        if path == "/api/match":
            self._handle_match(raw)
        elif path == "/api/export":
            self._handle_export(raw)
        elif path == "/api/xlsx/run":
            self._handle_xlsx(raw)
        else:
            self._send(404, "not found")

    def _handle_match(self, raw_body):
        try:
            payload = json.loads(raw_body.decode("utf-8") or "{}")
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            self._send_json({"error": "bad JSON: %s" % e}, 400)
            return

        request = {
            "palette": payload.get("palette") or DEFAULT_PALETTE,
            "targets": payload.get("targets") or [],
            "format": payload.get("format") or "hex",
            "rgb_scale": payload.get("rgb_scale") or "auto",
            "min_percent": payload.get("min_percent") if payload.get("min_percent") is not None else 0,
            "exclude": payload.get("exclude") or "",
        }
        if len(request["targets"]) == 0:
            self._send_json({"error": "no target rows", "rows": []}, 400)
            return

        worker = get_worker()
        if worker is None:
            self._send_json({"error": "CLI binary not found under build/. Run build.sh first."}, 500)
            return

        result = worker.query(request)
        if "error" in result and "rows" not in result:
            self._send_json(result, 400)
        else:
            self._send_json(result)

    def _handle_export(self, raw_body):
        """Export uses CSV columns. We run the same query and reformat to CSV
        client-side in the browser, so here we just return the JSON match
        result with a CSV content-type hint. Simpler: return JSON and let JS
        build CSV. But to keep one code path, we reuse /api/match logic."""
        # The browser's export already calls /api/match and builds CSV in JS
        # (see index.html exportCsv). This endpoint is kept for compatibility
        # but delegates to the same worker.
        self._handle_match(raw_body)

    def _handle_xlsx(self, raw_body):
        """Upload a .xlsx, batch-match its Lab targets with the current palette,
        write FS/Prusa results into new sheets, return the modified .xlsx.

        Request is multipart/form-data OR JSON with base64 file content.
        We accept JSON for simplicity: {xlsx: "<base64>", palette: [...],
        min_percent: N, exclude: "..."}.
        """
        try:
            import base64
            payload = json.loads(raw_body.decode("utf-8") or "{}")
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            self._send_json({"error": "bad request: %s" % e}, 400)
            return

        xlsx_b64 = payload.get("xlsx")
        if not xlsx_b64:
            self._send_json({"error": "missing 'xlsx' field (base64-encoded file)"}, 400)
            return
        try:
            xlsx_bytes = base64.b64decode(xlsx_b64)
        except Exception as e:
            self._send_json({"error": "bad base64: %s" % e}, 400)
            return

        palette = payload.get("palette") or DEFAULT_PALETTE
        min_percent = payload.get("min_percent")
        exclude = payload.get("exclude") or ""

        try:
            data, count = process_xlsx(xlsx_bytes, palette, min_percent, exclude)
        except Exception as e:
            self._send_json({"error": str(e)}, 500)
            return

        # Return the xlsx bytes directly for download.
        self._send(200, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   extra_headers={"Content-Disposition": 'attachment; filename="results.xlsx"'})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def open_browser_async(url):
    def _open():
        for cmd in (["start", "", url], ["xdg-open", url], ["open", url]):
            try:
                subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return
            except OSError:
                continue
    threading.Thread(target=_open, daemon=True).start()


def _port_in_use(port):
    """Quick check whether a TCP port is already bound."""
    import socket as _sock
    s = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
    s.settimeout(0.3)
    try:
        s.bind(("127.0.0.1", port))
        s.close()
        return False
    except OSError:
        return True


def main():
    cli = find_cli()
    if cli is None:
        print("[warn] build/color_match_batch not found — build it first (build.sh).", file=sys.stderr)

    # Pre-warm the CLI worker NOW (not on first request) so the first match
    # is just as fast as subsequent ones.
    if cli:
        try:
            get_worker()
            print("[ok] CLI worker ready")
        except Exception as e:
            print("[warn] CLI worker failed to start: %s" % e, file=sys.stderr)

    # Find a free port: try DEFAULT_PORT, then increment up to +20.
    socketserver.TCPServer.allow_reuse_address = True
    port = DEFAULT_PORT
    httpd = None
    for attempt in range(20):
        if _port_in_use(port):
            print("[info] port %d busy, trying %d ..." % (port, port + 1))
            port += 1
            continue
        try:
            httpd = socketserver.TCPServer(("127.0.0.1", port), Handler)
            break
        except OSError:
            port += 1
    if httpd is None:
        print("[error] could not find a free port in %d-%d" % (DEFAULT_PORT, port), file=sys.stderr)
        sys.exit(1)

    def _cleanup():
        global _WORKER
        if _WORKER:
            _WORKER.shutdown()
        httpd.shutdown()
    atexit.register(_cleanup)

    url = "http://localhost:%d" % port
    print("[color-mixer-batch] serving on %s" % url)
    print("  (Ctrl+C to stop)")
    # Flush so the user sees the URL before the browser opens.
    sys.stdout.flush()
    threading.Timer(0.5, lambda: open_browser_async(url)).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n[closed]")


if __name__ == "__main__":
    main()
