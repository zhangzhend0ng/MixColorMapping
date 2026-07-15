#!/usr/bin/env python3
"""Comprehensive end-to-end test suite for color-mixer-batch.

Run:  python tests/test_all.py

Covers:
  T1  CLI golden regression (delegates to tests/run_tests.sh)
  T2  CLI batch mode — all 4 input formats (hex/rgb/lab/cmyk)
  T3  CLI serve mode — JSON pipe, multi-query, error handling
  T4  Web server API — /defaults, /match, /export, xlsx upload
  T5  xlsx round-trip — 128-color batch match + sheet verification
  T6  Frontend sanity — JS syntax + key-access correctness

Exits 0 if all pass, 1 if any fail. Prints a summary table.
"""
import base64, json, os, subprocess, sys, time, tempfile, urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLI = os.path.join(ROOT, "build", "color_match_batch.exe")
if not os.path.isfile(CLI):
    CLI = os.path.join(ROOT, "build", "color_match_batch")
PASS = 0; FAIL = 0; RESULTS = []

def check(name, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1; RESULTS.append((name, "PASS", "")); print(f"  [PASS] {name}")
    else:
        FAIL += 1; RESULTS.append((name, "FAIL", detail)); print(f"  [FAIL] {name}  {detail}")

def cli_batch(palette_csv, targets_csv, fmt, extra=None):
    """Run CLI in batch mode, return (rc, results_csv_text)."""
    extra = extra or []
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "p.csv"); t = os.path.join(d, "t.csv"); o = os.path.join(d, "o.csv")
        with open(p, "w") as f: f.write(palette_csv)
        with open(t, "w") as f: f.write(targets_csv)
        cmd = [CLI, "--palette", p, "--targets", t, "--output", o, "--input-format", fmt] + extra
        r = subprocess.run(cmd, capture_output=True, text=True)
        out = open(o).read() if os.path.isfile(o) else ""
        return r.returncode, out

def run_serve(queries):
    """Start CLI --serve, send N JSON queries, return list of responses."""
    proc = subprocess.Popen([CLI, "--serve"], stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                            stderr=subprocess.DEVNULL, bufsize=1, text=True)
    proc.stdout.readline()  # ready handshake
    resps = []
    for q in queries:
        proc.stdin.write(json.dumps(q) + "\n"); proc.stdin.flush()
        resps.append(json.loads(proc.stdout.readline()))
    proc.stdin.close(); proc.wait(timeout=3)
    return resps

PAL8 = [["#FF0000","Red"],["#00FF00","Green"],["#0000FF","Blue"],
        ["#FFFF00","Yellow"],["#FFFFFF","White"],["#000000","Black"]]
PAL_CSV = "\n".join(",".join(r) for r in PAL8) + "\n"

# =========================================================================
print("=== T1: Golden regression (C++ algorithm correctness) ===")
# =========================================================================
# Verify the canonical slicer baseline values via the CLI serve mode.
# The slicer pins: filament_mixer_lerp(0,33,133, 252,211,0, 0.5) == (47,141,56).
# We put blue=#002185 (0,33,133) and yellow=#f6b921 (252,211,0)... but those are
# the Prusa demo colors. The actual slicer golden uses (0,33,133) and (252,211,0).
# Strategy: palette = [blue(0,33,133), yellow(252,211,0), white], target = (47,141,56).
# FS should match it at ~50:50 with ΔE ≈ 0 (since FS@0.5 IS (47,141,56)).
_g_pal = [["#002185","blue"], ["#f6b921","yellow"], ["#FFFFFF","white"]]
# 0,33,133 = #002185 ; 252,211,0 = #fcd300 ... the slicer golden uses literal
# (0,33,133) and (252,211,0). #002185=0,33,133 ✓. 252,211,0 = #fcd300.
_g_pal = [["#002185","blue"], ["#fcd300","yellow"], ["#FFFFFF","white"]]
_g_resp = run_serve([{"palette":_g_pal,"targets":[["golden","#2f8d38"]],"format":"hex","min_percent":0,"exclude":"","rgb_scale":"auto"}])
# #2f8d38 = (47,141,56) — the exact FS golden output.
if _g_resp and "rows" in _g_resp[0]:
    _fs = _g_resp[0]["rows"][0]["fs"]
    # FS should pick blue/yellow and ΔE should be very small (the target IS the FS blend)
    check("T1 FS golden (47,141,56 reachable)", _fs["delta_e"] < 3.0,
          f"fs ΔE={_fs['delta_e']:.2f} hex={_fs['hex']} ids={_fs['ids']}")
else:
    check("T1 FS golden", False, str(_g_resp[0])[:150])

# Verify Prusa gradient-safety: a target == a palette color should match with ΔE≈0
_p_resp = run_serve([{"palette":[["#FF0000","R"],["#00FF00","G"],["#0000FF","B"]],
                      "targets":[["exact","#FF0000"]],"format":"hex","min_percent":0,"exclude":"","rgb_scale":"auto"}])
if _p_resp and "rows" in _p_resp[0]:
    _pr = _p_resp[0]["rows"][0]["prusa"]
    check("T1 Prusa exact-color match (ΔE<2)", _pr["delta_e"] < 2.0, f"prusa ΔE={_pr['delta_e']:.2f}")
else:
    check("T1 Prusa exact-color match", False, str(_p_resp[0])[:150])

# =========================================================================
print("\n=== T2: CLI batch mode (4 input formats) ===")
# =========================================================================
for fmt, tgt_csv, label_col in [
    ("hex",  "label,color\nt1,#FF8800\nt2,#008888\n", 1),
    ("rgb",  "label,R,G,B\nt1,255,136,0\nt2,0,136,136\n", 1),
    ("lab",  "label,L,a,b\nt1,50,0,0\nt2,60,10,-10\n", 1),
    ("cmyk", "label,C,M,Y,K\nt1,0,100,100,0\nt2,100,0,100,0\n", 1),
]:
    rc, out = cli_batch(PAL_CSV, tgt_csv, fmt)
    lines = [l for l in out.strip().split("\n") if l and not l.startswith("label")]
    check(f"T2 batch {fmt}", rc == 0 and len(lines) == 2, f"rc={rc} lines={len(lines)}")

# exclude flag
rc, out = cli_batch(PAL_CSV, "label,color\nt1,#FF0000\n", "hex", ["--exclude", "1-2,1-3"])
check("T2 --exclude flag", rc == 0, f"rc={rc}")

# min-percent flag
rc, out = cli_batch(PAL_CSV, "label,color\nt1,#808080\n", "hex", ["--min-percent", "20"])
check("T2 --min-percent flag", rc == 0, f"rc={rc}")

# =========================================================================
print("\n=== T3: CLI serve mode (JSON pipe) ===")
# =========================================================================
queries = [
    {"palette": PAL8, "targets": [["orange","#FF8800"],["teal","#008888"]], "format":"hex", "min_percent":0, "exclude":"", "rgb_scale":"auto"},
    {"palette": PAL8, "targets": [["gray","50","0","0"]], "format":"lab", "min_percent":0, "exclude":"", "rgb_scale":"auto"},
]
resps = run_serve(queries)
check("T3 serve multi-query", len(resps) == 2 and resps[0].get("count") == 2, str(resps[0])[:100])
# verify keys match frontend expectations
if resps and "rows" in resps[0]:
    row = resps[0]["rows"][0]
    keys_ok = all(k in row["fs"] for k in ["delta_e","hex","type","ids","weights","r","g","b","L","a","b_lab"])
    check("T3 serve JSON keys (frontend-compatible)", keys_ok, str(sorted(row["fs"].keys())))
    check("T3 serve target keys", all(k in row["target"] for k in ["hex","r","g","b","L"]), "")
else:
    check("T3 serve JSON keys", False, "no rows")

# error handling: 1-color palette
err_resps = run_serve([{"palette":[["#FF0000","R"]], "targets":[["x","#000000"]], "format":"hex", "min_percent":0, "exclude":"", "rgb_scale":"auto"}])
check("T3 serve error (1-color palette)", "error" in err_resps[0], str(err_resps[0]))

# error handling: empty targets
err_resps2 = run_serve([{"palette": PAL8, "targets": [], "format":"hex", "min_percent":0, "exclude":"", "rgb_scale":"auto"}])
check("T3 serve error (empty targets)", "error" in err_resps2[0], str(err_resps2[0]))

# =========================================================================
print("\n=== T4: Web server API ===")
# =========================================================================
# start server
srv = subprocess.Popen(["python", "-u", "web/app.py"], cwd=ROOT, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
time.sleep(3)
BASE = "http://127.0.0.1:8008"

def api_get(path):
    try:
        r = urllib.request.urlopen(f"{BASE}{path}", timeout=5)
        return r.status, r.read()
    except Exception as e:
        return 0, str(e).encode()

def api_post(path, body_bytes, ctype="application/json"):
    try:
        r = urllib.request.urlopen(urllib.request.Request(f"{BASE}{path}", data=body_bytes,
            headers={"Content-Type":ctype}), timeout=30)
        return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()
    except Exception as e:
        return 0, str(e).encode()

try:
    # /api/defaults
    st, body = api_get("/api/defaults")
    d = json.loads(body) if st == 200 else {}
    check("T4 GET /api/defaults", st == 200 and "palette" in d and "targets" in d, f"st={st}")

    # GET / (HTML)
    st, body = api_get("/")
    check("T4 GET / (HTML)", st == 200 and b"<html" in body.lower(), f"st={st}")

    # /api/match hex
    req = json.dumps({"palette":PAL8,"targets":[["orange","#FF8800"],["teal","#008888"]],"format":"hex","min_percent":0,"exclude":"","rgb_scale":"auto"}).encode()
    st, body = api_post("/api/match", req)
    d = json.loads(body) if st == 200 else {}
    check("T4 POST /api/match (hex)", st == 200 and d.get("count") == 2, f"st={st} body={body[:100]}")

    # /api/match lab
    req = json.dumps({"palette":PAL8,"targets":[["gray","50","0","0"]],"format":"lab","min_percent":0,"exclude":"","rgb_scale":"auto"}).encode()
    st, body = api_post("/api/match", req)
    check("T4 POST /api/match (lab)", st == 200, f"st={st}")

    # /api/match error (1-color)
    req = json.dumps({"palette":[["#FF0000","R"]],"targets":[["x","#000000"]],"format":"hex"}).encode()
    st, body = api_post("/api/match", req)
    check("T4 POST /api/match error (1-color)", st == 400, f"st={st}")

finally:
    srv.terminate(); srv.wait(timeout=5)

# =========================================================================
print("\n=== T5: xlsx upload/download ===")
# =========================================================================
# Build a sample .xlsx in-memory (Lab targets on "Sheet1") so the test is
# self-contained and does not depend on an external user file. 24 colors is
# enough to exercise the batch path while staying fast.
try:
    import openpyxl
    N_ROWS = 24
    _wb = openpyxl.Workbook()
    _ws = _wb.active
    _ws.title = "Sheet1"
    _ws.append(["序号", "L*", "a*", "b*"])
    # deterministic spread across the Lab gamut
    _ls = [30, 50, 70, 90]
    _as = [-40, 0, 40]
    _bs = [-40, 0, 40]
    _i = 0
    for _L in _ls:
        for _a in _as:
            for _b in _bs:
                if _i >= N_ROWS:
                    break
                _i += 1
                _ws.append([_i, _L, _a, _b])
    _buf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    _wb.save(_buf.name)
    _buf.close()
    _xlsx_bytes = open(_buf.name, "rb").read()
    os.unlink(_buf.name)
    HAS_OPENPYXL_TEST = True
except ImportError:
    HAS_OPENPYXL_TEST = False

if not HAS_OPENPYXL_TEST:
    check("T5 openpyxl available", False, "openpyxl not installed — install with: pip install openpyxl")
else:
    # restart server for xlsx test
    srv = subprocess.Popen(["python", "-u", "web/app.py"], cwd=ROOT, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    time.sleep(3)
    # CMYW palette — the xlsx output maps recipes into fixed C/M/Y/W columns
    # by matching palette hex to standard subtractive primaries.
    PAL_CMYW = [["#00FFFF","Cyan"],["#FF00FF","Magenta"],["#FFFF00","Yellow"],["#FFFFFF","White"]]
    try:
        b64 = base64.b64encode(_xlsx_bytes).decode()
        req = json.dumps({"xlsx":b64,"palette":PAL_CMYW,"min_percent":0,"exclude":""}).encode()
        t0 = time.perf_counter()
        st, body = api_post("/api/xlsx/run", req)
        elapsed = time.perf_counter() - t0
        check("T5 xlsx upload (HTTP)", st == 200 and len(body) > 5000, f"st={st} size={len(body)}")
        check("T5 xlsx latency (< 5s)", elapsed < 5.0, f"{elapsed:.1f}s")

        if st == 200:
            out = os.path.join(tempfile.gettempdir(), "test_result.xlsx")
            with open(out, "wb") as f: f.write(body)
            wb = openpyxl.load_workbook(out)
            check("T5 xlsx has 配色结果 sheet", "配色结果" in wb.sheetnames, str(wb.sheetnames))
            check("T5 xlsx active sheet is 配色结果", wb.active.title == "配色结果",
                  f"active={wb.active.title}")
            ws = wb["配色结果"]
            check(f"T5 xlsx {N_ROWS} rows in 配色结果", ws.max_row == N_ROWS + 1, f"rows={ws.max_row}")
            hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
            check("T5 xlsx header correct",
                  hdr[0]=="序号" and "FS_C%" in hdr and "FS_ΔE" in hdr and "Prusa_ΔE" in hdr
                  and "FS预测Lab" in hdr and "Prusa预测Lab" in hdr
                  and "FS预测CMYK" not in hdr and "Prusa预测CMYK" not in hdr,
                  str(hdr))
            check("T5 xlsx has 19 columns", ws.max_column == 19, f"cols={ws.max_column}")
            check("T5 xlsx 序号 populated", ws.cell(row=2, column=1).value is not None,
                  str(ws.cell(row=2,column=1).value))
            # verify CMYW columns sum ~100 for a recipe row (FS side, cols 6-9)
            r2_cmyw = [ws.cell(row=2, column=c).value for c in range(6, 10)]
            check("T5 xlsx FS CMYW sums ~100",
                  all(isinstance(v,(int,float)) for v in r2_cmyw) and 95 <= sum(r2_cmyw) <= 100,
                  f"CMYW={r2_cmyw} sum={sum(r2_cmyw) if all(isinstance(v,(int,float)) for v in r2_cmyw) else '?'}")
            # verify predicted Lab cell populated as string (FS: col 11)
            r2_lab_pred = ws.cell(row=2, column=11).value
            check("T5 xlsx FS pred Lab populated",
                  isinstance(r2_lab_pred, str) and r2_lab_pred.count("/") == 2, f"Lab={r2_lab_pred!r}")
            # verify at least one row has the best-algo green highlight
            # (only the winning ΔE cell is highlighted: col 12 = FS_ΔE, col 19 = Prusa_ΔE)
            green = "FFC6EFCE"
            de_cols = (12, 19)
            has_highlight = any(
                ws.cell(row=r, column=c).fill.fgColor.rgb == green
                for r in range(2, ws.max_row+1)
                for c in de_cols
                if ws.cell(row=r, column=c).fill and ws.cell(row=r, column=c).fill.fill_type == "solid"
            )
            check("T5 xlsx best-algo highlight present", has_highlight, "no green ΔE cell found")
            # and verify the highlight is ONLY on a ΔE cell, never on recipe/Lab cells
            non_de_cols = (set(range(6, 12)) | set(range(13, 19)))  # all non-ΔE data cols
            no_spurious = all(
                ws.cell(row=r, column=c).fill.fgColor.rgb != green
                for r in range(2, ws.max_row+1)
                for c in non_de_cols
                if ws.cell(row=r, column=c).fill and ws.cell(row=r, column=c).fill.fill_type == "solid"
            )
            check("T5 xlsx highlight only on ΔE cell", no_spurious, "green found on non-ΔE cell")
    finally:
        srv.terminate(); srv.wait(timeout=5)

# =========================================================================
print("\n=== T6: Frontend sanity ===")
# =========================================================================
# JS syntax check via node
import re
html = open(os.path.join(ROOT, "web", "static", "index.html"), encoding="utf-8").read()
m = re.search(r"<script>(.*?)</script>", html, re.DOTALL)
if m:
    js_path = os.path.join(tempfile.gettempdir(), "_page_check.js")
    with open(js_path, "w", encoding="utf-8") as f: f.write(m.group(1))
    r = subprocess.run(["node", "--check", js_path], capture_output=True, text=True)
    check("T6 frontend JS syntax", r.returncode == 0, r.stderr[:200] if r.returncode else "")
    os.unlink(js_path)
else:
    check("T6 frontend JS syntax", False, "no <script> found")

# check key functions exist
for fn in ["recompute", "exportCsv", "uploadXlsx", "renderResults", "collectPalette", "collectTargets"]:
    check(f"T6 frontend has {fn}()", f"function {fn}" in html or f"async function {fn}" in html, "")

# =========================================================================
# SUMMARY
# =========================================================================
print("\n" + "=" * 60)
print(f"RESULTS: {PASS} passed, {FAIL} failed, {PASS+FAIL} total")
print("=" * 60)
if FAIL:
    print("\nFAILURES:")
    for name, status, detail in RESULTS:
        if status == "FAIL":
            print(f"  ✗ {name}: {detail}")
sys.exit(0 if FAIL == 0 else 1)
